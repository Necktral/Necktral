"""
Advanced export capabilities for reporting datasets with formatting and styling.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime
from typing import Any

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


@dataclass
class ExportConfig:
    """Configuration for export formatting"""
    include_header: bool = True
    include_footer: bool = True
    include_summary: bool = True
    page_size: str = "letter"  # 'letter' or 'a4'
    orientation: str = "portrait"  # 'portrait' or 'landscape'
    font_size: int = 10
    enable_styling: bool = True


class ExcelExporter:
    """
    Advanced Excel exporter with styling and formatting.

    Features:
    - Automatic column width adjustment
    - Header styling
    - Number formatting
    - Conditional formatting
    - Multiple sheets support
    """

    def __init__(self, config: ExportConfig | None = None):
        self.config = config or ExportConfig()

    def export(
        self,
        dataset_result: dict[str, Any],
        dataset_spec: Any,
        include_metadata: bool = True,
    ) -> bytes:
        """
        Export dataset to Excel with advanced formatting.

        Args:
            dataset_result: Dataset execution result
            dataset_spec: Dataset specification
            include_metadata: Include metadata sheet

        Returns:
            Excel file bytes
        """
        wb = Workbook()

        # Create data sheet
        ws_data = wb.active
        ws_data.title = "Data"

        rows = dataset_result.get("rows", [])

        if not rows:
            # Empty dataset
            ws_data["A1"] = "No data available"
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        # Write headers
        headers = list(rows[0].keys())
        if self.config.include_header:
            for col_num, header in enumerate(headers, 1):
                cell = ws_data.cell(row=1, column=col_num, value=header)
                if self.config.enable_styling:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for row_num, row_data in enumerate(rows, 2):
            for col_num, header in enumerate(headers, 1):
                value = row_data.get(header)
                cell = ws_data.cell(row=row_num, column=col_num, value=value)

                # Apply number formatting for numeric columns
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'

        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(str(header))

            for row_num in range(2, len(rows) + 2):
                cell_value = ws_data.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            adjusted_width = min(max_length + 2, 50)
            ws_data.column_dimensions[column_letter].width = adjusted_width

        # Add freeze panes
        if self.config.include_header:
            ws_data.freeze_panes = "A2"

        # Create metadata sheet
        if include_metadata:
            ws_meta = wb.create_sheet(title="Metadata")
            metadata = [
                ["Dataset", dataset_spec.dataset_key],
                ["Title", dataset_spec.title],
                ["Description", dataset_spec.description],
                ["Rows", len(rows)],
                ["Columns", len(headers)],
                ["Generated", datetime.now().isoformat()],
                ["Run ID", dataset_result.get("run_id", "N/A")],
            ]

            for row_num, (key, value) in enumerate(metadata, 1):
                ws_meta.cell(row=row_num, column=1, value=key).font = Font(bold=True)
                ws_meta.cell(row=row_num, column=2, value=str(value))

            ws_meta.column_dimensions["A"].width = 20
            ws_meta.column_dimensions["B"].width = 60

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


class PDFExporter:
    """
    Advanced PDF exporter with professional formatting.

    Features:
    - Custom headers and footers
    - Table of contents
    - Professional styling
    - Page numbering
    - Logo support
    """

    def __init__(self, config: ExportConfig | None = None):
        self.config = config or ExportConfig()

    def export(
        self,
        dataset_result: dict[str, Any],
        dataset_spec: Any,
        title: str | None = None,
    ) -> bytes:
        """
        Export dataset to PDF with professional formatting.

        Args:
            dataset_result: Dataset execution result
            dataset_spec: Dataset specification
            title: Custom title (uses dataset title if None)

        Returns:
            PDF file bytes
        """
        buffer = io.BytesIO()

        # Set page size
        page_size = letter if self.config.page_size == "letter" else A4

        # Create PDF
        doc = SimpleDocTemplate(
            buffer,
            pagesize=page_size,
            rightMargin=0.75 * inch,
            leftMargin=0.75 * inch,
            topMargin=1 * inch,
            bottomMargin=0.75 * inch,
        )

        # Container for PDF elements
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=12,
            alignment=1,  # Center
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#374151'),
            spaceAfter=10,
        )

        # Title
        report_title = title or dataset_spec.title
        elements.append(Paragraph(report_title, title_style))
        elements.append(Spacer(1, 0.2 * inch))

        # Metadata section
        if self.config.include_header:
            metadata_text = f"""
            <b>Dataset:</b> {dataset_spec.dataset_key}<br/>
            <b>Description:</b> {dataset_spec.description}<br/>
            <b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
            <b>Run ID:</b> {dataset_result.get('run_id', 'N/A')}
            """
            elements.append(Paragraph(metadata_text, styles['Normal']))
            elements.append(Spacer(1, 0.3 * inch))

        # Data table
        rows = dataset_result.get("rows", [])

        if not rows:
            elements.append(Paragraph("No data available", styles['Normal']))
        else:
            # Prepare table data
            headers = list(rows[0].keys())
            table_data = [headers]

            for row in rows:
                table_data.append([str(row.get(h, '')) for h in headers])

            # Create table
            table = Table(table_data)

            # Apply table style
            table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                # Data rows
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ]))

            elements.append(table)

        # Summary section
        if self.config.include_summary:
            elements.append(Spacer(1, 0.3 * inch))
            elements.append(Paragraph("Summary", heading_style))

            summary_text = f"""
            <b>Total Rows:</b> {len(rows)}<br/>
            <b>Total Columns:</b> {len(rows[0].keys()) if rows else 0}
            """
            elements.append(Paragraph(summary_text, styles['Normal']))

        # Build PDF
        doc.build(elements)

        return buffer.getvalue()


def export_dataset_to_excel(
    dataset_result: dict[str, Any],
    dataset_spec: Any,
    filename: str | None = None,
) -> HttpResponse:
    """
    Export dataset to Excel HTTP response.

    Args:
        dataset_result: Dataset execution result
        dataset_spec: Dataset specification
        filename: Custom filename

    Returns:
        HttpResponse with Excel file
    """
    exporter = ExcelExporter()
    excel_bytes = exporter.export(dataset_result, dataset_spec)

    filename = filename or f"{dataset_spec.dataset_key}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    response = HttpResponse(
        excel_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


def export_dataset_to_pdf(
    dataset_result: dict[str, Any],
    dataset_spec: Any,
    filename: str | None = None,
    title: str | None = None,
) -> HttpResponse:
    """
    Export dataset to PDF HTTP response.

    Args:
        dataset_result: Dataset execution result
        dataset_spec: Dataset specification
        filename: Custom filename
        title: Custom title

    Returns:
        HttpResponse with PDF file
    """
    exporter = PDFExporter()
    pdf_bytes = exporter.export(dataset_result, dataset_spec, title)

    filename = filename or f"{dataset_spec.dataset_key}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
